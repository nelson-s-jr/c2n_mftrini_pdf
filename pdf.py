import PyPDF2
import os
import openpyxl
from time import time
from math import ceil


def print_process_time(tempo, num_fichas):
    tempo_medio_por_ficha = tempo / num_fichas
    
    if tempo < 60:
        print(f"Tempo Total de processamento: {round(tempo, 2)} segundos")
    else:
        minutos = int(tempo // 60)
        segundos = ceil(tempo % 60)
        print(f"Tempo de processamento: {minutos} minutos e {segundos} segundos.")
    
    print(f"Tempo médio por ficha: {round(tempo_medio_por_ficha, 2)} segundos.\n")




try:
    
    pasta_pdf = str(input('Informe o caminho dos arquivos pdf: ')).strip()
    
    caminho_planilha = str(input('Informe o caminho da planilha: ')).strip()
    
    if '"' in caminho_planilha:
        caminho_planilha = caminho_planilha.strip('"')
    if '"' in pasta_pdf:
        pasta_pdf = pasta_pdf.strip('"')
    
    
    inicio = time()
    resultados = {}
    contador = 0
    
    for nome_arquivo in os.listdir(pasta_pdf):
        # Verifica se o arquivo é um arquivo PDF
        if nome_arquivo.endswith('.pdf'):
            contador += 1
            # Caminho completo para o arquivo PDF
            caminho_arquivo = os.path.join(pasta_pdf, nome_arquivo)
            
            # Abre o arquivo PDF no modo de leitura binária ('rb')
            with open(caminho_arquivo, 'rb') as file:
                
                # Cria um objeto PdfReader
                pdf_reader = PyPDF2.PdfReader(file)
                
                page = pdf_reader.pages[0]
                text = page.extract_text()
                # Procura a posição da palavra "POSITIVE" no texto
                
                ficha = nome_arquivo.split('.')[0]
                
                posicao_positive = text.find("POSITIVE")
                if posicao_positive != -1:
                    resultados[ficha] = 'POSITIVO'
                    print(f"O resultado da ficha {ficha} é POSITIVO.")
                
                # Procura a posição da palavra "NEGATIVE" no texto
                posicao_negative = text.find("NEGATIVE")
                if posicao_negative != -1:
                    print(f"O resultado da ficha {ficha} é NEGATIVO.")
                    resultados[ficha] = 'NEGATIVO'
    
    wb = openpyxl.load_workbook(caminho_planilha, keep_vba=True)
    
    # Seleciona a aba na planilha
    sheet = wb['DADOS']
    
    # Encontre o índice das colunas 'FICHA' e 'RESULTADO'
    index_ficha = None
    index_resultado = None
    for idx, cell in enumerate(sheet[1], 1):  # Iterar sobre a primeira linha (cabeçalho)
        if cell.value == 'FICHA':
            index_ficha = idx
        elif cell.value == 'RESULTADO':
            index_resultado = idx
    
    if index_ficha is not None and index_resultado is not None:
        fichas_encontradas = 0
        # Iterar sobre as linhas para preencher 'RESULTADO' com base no dicionário
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1,
                                   max_col=max(index_ficha, index_resultado)):
            chave = str(row[index_ficha - 1].value) if row[
                index_ficha - 1].value else ''  # Convertendo para string e tratando valores vazios
            if chave in resultados:
                row[index_resultado - 1].value = resultados[chave]
                fichas_encontradas += 1
                print(f"Ficha '{chave}' encontrada e atualizada com o resultado '{resultados[chave]}'.")
    
    if fichas_encontradas == 0:
        print("Nenhuma das fichas foi encontrada na planilha.")
    
    # Salvar a planilha atualizada
    print('\nSalvando a planilha...\n')
    wb.save(caminho_planilha)
    print('Alterações salvas com sucesso.\n')
    
    fim = time()
    
    tempo_total = fim - inicio
    
    print_process_time(tempo_total, contador)

except Exception as erro:
    print(f'Houve o erro {erro}')

input('Pressione ENTER para encerrar o programa')