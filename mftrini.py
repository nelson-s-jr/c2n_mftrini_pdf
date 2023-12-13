import PyPDF2
import os
from openpyxl import load_workbook
from time import time
from math import ceil
import shutil
from tkinter import filedialog

print("Selecione os arquivos para preencher a planilha do MFTRINI-AMP")

def backup_excel(caminho_planilha):
    nome_backup = os.path.basename(caminho_planilha).replace(".xlsm", "_BACKUP.xlsm")
    diretorio_backup = os.path.join(os.path.dirname(caminho_planilha), "BACKUP")
    caminho_backup = os.path.join(diretorio_backup, nome_backup)
    if not os.path.isdir(diretorio_backup):
        os.mkdir(diretorio_backup)
        print("Pasta BACKUP criada.")
    else:
        print("Pasta BACKUP ja existente.")
    shutil.copy2(caminho_planilha, caminho_backup)


def print_process_time(tempo, num_fichas):
    tempo_medio_por_ficha = tempo / num_fichas
    if tempo < 60:
        print(f"Tempo Total de processamento: {round(tempo, 2)} segundos")
    else:
        minutos = int(tempo // 60)
        segundos = ceil(tempo % 60)
        print(f"Tempo de processamento: {minutos} minutos e {segundos} segundos.")  
    print(f"Tempo médio por ficha: {round(tempo_medio_por_ficha, 2)} segundos.\n")


# Função para extrair a fração de DNA fetal
def extrair_fracao_dna(texto):
    inicio = texto.find('fração de dna fetal')
    if inicio != -1:
        fracao = text[text.find('fração de dna fetal'): text.find('%') + 1].split(':')[1].strip()
        return str(fracao)
    return None


try:
    pasta_pdf = filedialog.askdirectory(title="Escolha a pasta com os resultados do MFTRINI-AMP a serem salvos")
    pasta_pdf = rf"{pasta_pdf}"
    
    caminho_planilha = filedialog.askopenfilename(title="Selecione a planilha MFTRINI-AMP a ser preenchida")
    caminho_planilha = rf"{caminho_planilha}"
    backup_excel(caminho_planilha)

    contador_pdf = 0
    resultados = {}
    
    doencas = [
        'trissomia 21: alto risco','trissomia 21: risco aumentado' 'trissomia 18: alto risco', 'trissomia 18: risco aumentado','trissomia 13: alto risco', 
        'trissomia 13: risco aumentado','monossomia x: alto risco', 'monossomia x: risco aumentado','triploidia: alto risco', 'triploidia: risco aumentado', 
        'deleção 22q 11.2: alto risco','deleção 22q 11.2: risco aumentado','deleção 1p36: alto risco',
        'deleção 1p36: risco aumentado', 'síndrome de angelman: alto risco', 'síndrome de angelman: risco aumentado', 
        'síndrome cri-du-chat: alto risco', 'síndrome cri-du-chat: risco aumentado','síndrome prader-willi: alto risco'
        'síndrome prader-willi: risco aumentado'
    ]
    
    print('\nVerificando PDFs\n')
    
    for nome_arquivo in os.listdir(pasta_pdf):
        if nome_arquivo.endswith('.pdf'):
            contador_pdf += 1
            caminho_arquivo = os.path.join(pasta_pdf, nome_arquivo)
            doencas_alto_risco = []
            ficha = nome_arquivo.split('.pdf')[0]
            
            with open(caminho_arquivo, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                num_pages = len(pdf_reader.pages)
                
                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    text = page.extract_text().lower()
                    
                    lista = []
                    
                    if 'resultado: alto risco' in text:
                        status = 'Alto Risco'
                        fracao_dna = extrair_fracao_dna(text)
                    
                    if any(doenca in text for doenca in doencas):
                        for doenca in doencas:
                            if doenca in text:
                                nome_doenca = doenca.split(':')[0].strip().title()
                                doencas_alto_risco.append(nome_doenca)

                                if len(doencas_alto_risco) == 0:
                                    doencas_alto_risco.append("Sem resultado")
                                
                                if len(doencas_alto_risco) == 1:
                                    print(f'Ficha: {ficha}\nResultado: Alto Risco\nFração de DNA Fetal: {fracao_dna}')
                                    print(f'Doença Encontrada: {nome_doenca}\n')
                                elif len(doencas_alto_risco) >= 2:
                                    print(f'Ficha: {ficha}\nResultado: Alto Risco\nFração de DNA Fetal: {fracao_dna}')
                                    print(f"Doenças Encontradas: {', '.join(doencas_alto_risco)}\n")
                                



                        lista.extend([status, ', '.join(doencas_alto_risco) if doencas_alto_risco else '-', fracao_dna])
                        resultados[ficha] = lista

                    
                    elif 'resultado: baixo risco' in text:
                        status = 'Baixo Risco'
                        fracao_dna = extrair_fracao_dna(text)
                        
                        print(f'Ficha: {ficha}\nResultado: {status}\nFração de DNA Fetal: {fracao_dna}\n')
                        
                        lista.extend([status, '-', fracao_dna])
                        resultados[ficha] = lista
                    

                    elif any(
                        resultado in text for resultado in ['resultado: achado atípico', 'resultado: achado atipico']):
                        status = 'Achado Atípico'
                        fracao_dna = extrair_fracao_dna(text)
                        
                        print(f'Ficha: {ficha}\nResultado: {status}\nFração de DNA Fetal: {fracao_dna}\n')
                        
                        lista.extend([status, '-', fracao_dna])
                        resultados[ficha] = lista
                    
                    if 'resultado: vide nota' in text:
                        status = 'Verificar manual'
                        fracao_dna = extrair_fracao_dna(text)
                        
                        print(f'Ficha: {ficha}\nResultado: {status}\nFração de DNA Fetal: {fracao_dna}\n')
                        
                        lista.extend([status, '-', fracao_dna])
                        resultados[ficha] = lista
    
    if contador_pdf == 0:
        print('Nenhuma ficha encontrada.\n')
    elif contador_pdf == 1:
        print(f"Total de {contador_pdf} ficha encontrada.\n")
    else:
        print(f"Total de {contador_pdf} fichas encontradas.\n")
    
    # Carregar o arquivo Excel existente
    workbook = load_workbook(caminho_planilha, keep_vba=True)
    inicio = time()
    print('Acessando a planilha...\n')
    
    # Selecionar a planilha na qual você quer trabalhar
    sheet = workbook['MFTRINIAMP']  # Ou utilize workbook['nome_da_planilha'] se não for a planilha ativa
    
    # Encontrar os cabeçalhos 'FICHA', 'RESULTADO', 'ANEUPLOIDIAS' e 'FRAÇÃO fetal' na primeira linha
    headers = {'FICHA': None, 'RESULTADO': None, 'ANEUPLOIDIAS': None, 'FRAÇÃO FETAL': None}
    for col in sheet.iter_cols(min_row=1, max_row=1):
        if col[0].value in headers:
            headers[col[0].value] = col[0].column_letter
    
    # Procurar cada número da ficha na coluna 'FICHA' e escrever os valores correspondentes nas colunas especificadas
    ficha_column = headers['FICHA']
    result_column = headers['RESULTADO']
    aneuploidias_column = headers['ANEUPLOIDIAS']
    fracao_fetal_column = headers['FRAÇÃO FETAL']
    
    if ficha_column:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet[ficha_column][0].column,
                                   max_col=sheet.max_column):
            chave = str(row[0].value)  # Convertendo para string para garantir a correspondência
            if chave in resultados:
                print(f"Preenchendo as informações correspondentes da ficha {chave}")
                resultado, aneuploidias, fracao_fetal = resultados[chave]
                
                # Escrever os valores correspondentes nas colunas 'RESULTADO', 'ANEUPLOIDIAS' e 'FRAÇÃO fetal'
                sheet[f'{result_column}{row[0].row}'] = resultado
                sheet[f'{aneuploidias_column}{row[0].row}'] = aneuploidias
                sheet[f'{fracao_fetal_column}{row[0].row}'] = fracao_fetal
    
    # Salvar as alterações no arquivo Excel
    print('\nSalvando alterações na planilha\n')
    workbook.save(caminho_planilha)
    print('Alterações salvas com sucesso.\n')

except Exception as erro:
    print(f'Ocorreu o erro {erro}')

fim = time()

tempo_total = fim - inicio

print_process_time(tempo_total, contador_pdf)

input('Programa encerrado.')


