import PyPDF2
import os
import openpyxl
from time import time
from math import ceil
import shutil
from tkinter import filedialog

print("Selecione os arquivos para preencher a planilha da C2N")

def backup_excel(caminho_planilha):
    nome_backup = os.path.basename(caminho_planilha).replace(".xlsm", "_BACKUP.xlsm")
    diretorio_backup = os.path.join(os.path.dirname(caminho_planilha), "BACKUP")
    caminho_backup = os.path.join(diretorio_backup, nome_backup)
    if not os.path.isdir(diretorio_backup):
        os.mkdir(diretorio_backup)
        print("Pasta BACKUP criada.")
    else:
        print("Pasta BACKUP ja existente.")
    shutil.copy2(caminho_planilha, caminho_backup)



def print_process_time(tempo, num_fichas):
    tempo_medio_por_ficha = tempo / num_fichas
    
    if tempo < 60:
        print(f"Tempo Total de processamento: {round(tempo, 2)} segundos")
    else:
        minutos = int(tempo // 60)
        segundos = ceil(tempo % 60)
        print(f"Tempo de processamento: {minutos} minutos e {segundos} segundos.")
    
    print(f"Tempo médio por ficha: {round(tempo_medio_por_ficha, 2)} segundos.\n")




try:
        
    pasta_pdf = filedialog.askdirectory(title="Escolha a pasta com os resultados da C2N a serem salvos")
    pasta_pdf = rf"{pasta_pdf}"
    
    caminho_planilha = filedialog.askopenfilename(title="Selecione a planilha da C2N a ser preenchida")
    caminho_planilha = rf"{caminho_planilha}"

    
    backup_excel(caminho_planilha)

    inicio = time()
    resultados = {}
    contador = 0
    
    for nome_arquivo in os.listdir(pasta_pdf):
        # Verifica se o arquivo é um arquivo PDF
        if nome_arquivo.endswith('.pdf'):
            contador += 1
            # Caminho completo para o arquivo PDF
            caminho_arquivo = os.path.join(pasta_pdf, nome_arquivo)
            
            # Abre o arquivo PDF no modo de leitura binária ('rb')
            with open(caminho_arquivo, 'rb') as file:
                
                # Cria um objeto PdfReader
                pdf_reader = PyPDF2.PdfReader(file)
                
                page = pdf_reader.pages[0]
                text = page.extract_text()
                # Procura a posição da palavra "POSITIVE" no texto
                
                ficha = nome_arquivo.split('.')[0]
                
                posicao_positive = text.find("POSITIVE")
                if posicao_positive != -1:
                    resultados[ficha] = 'POSITIVO'
                    print(f"O resultado da ficha {ficha} é POSITIVO.")
                
                # Procura a posição da palavra "NEGATIVE" no texto
                posicao_negative = text.find("NEGATIVE")
                if posicao_negative != -1:
                    print(f"O resultado da ficha {ficha} é NEGATIVO.")
                    resultados[ficha] = 'NEGATIVO'

                if posicao_negative == -1 and posicao_positive == -1:
                    print(f"Verificar manualmente a ficha {ficha}.")
                    resultados[ficha] = 'VERIFICAR'
    
    wb = openpyxl.load_workbook(caminho_planilha, keep_vba=True)
    
    # Seleciona a aba na planilha
    sheet = wb['DADOS']
    
    # Encontre o índice das colunas 'FICHA' e 'RESULTADO'
    index_ficha = None
    index_resultado = None
    for idx, cell in enumerate(sheet[1], 1):  # Iterar sobre a primeira linha (cabeçalho)
        if cell.value == 'FICHA':
            index_ficha = idx
        elif cell.value == 'RESULTADO':
            index_resultado = idx
    
    if index_ficha is not None and index_resultado is not None:
        fichas_encontradas = 0
        # Iterar sobre as linhas para preencher 'RESULTADO' com base no dicionário
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1,
                                   max_col=max(index_ficha, index_resultado)):
            chave = str(row[index_ficha - 1].value) if row[
                index_ficha - 1].value else ''  # Convertendo para string e tratando valores vazios
            if chave in resultados:
                row[index_resultado - 1].value = resultados[chave]
                fichas_encontradas += 1
                print(f"Ficha '{chave}' encontrada e atualizada com o resultado '{resultados[chave]}'.")
    
    if fichas_encontradas == 0:
        print("Nenhuma das fichas foi encontrada na planilha.")
    
    # Salvar a planilha atualizada
    print('\nSalvando a planilha...\n')
    wb.save(caminho_planilha)
    print('Alterações salvas com sucesso.\n')
    
    fim = time()
    
    tempo_total = fim - inicio
    
    print_process_time(tempo_total, contador)

except Exception as erro:
    print(f'Houve o erro {erro}')

input('Programa encerrado.')

